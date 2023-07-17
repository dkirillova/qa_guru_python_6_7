import os
from openpyxl import load_workbook
from .conftest import resources_path


def test_read_excel_cell():
    file_path = os.path.join(resources_path, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(file_path)
    sheet = workbook.active
    cell_value = sheet.cell(row=3, column=2).value

    assert cell_value is not None, "Failed to read cell value"
    assert cell_value == "Mara"



